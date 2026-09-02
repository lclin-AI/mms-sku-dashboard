"""
Pull MMS Daily Order Reports and upsert SKU-level sales into Supabase.

Grain written: (store_code, order_date, delivery_date, sku_id) -> qty, amount, lines.
A re-run of a given order_date REPLACES exactly that day's rows for the store,
so it is safe to run repeatedly (today's file is a partial snapshot and will be
corrected on the next run).

MMS retention is only ~12 days, so run this daily; Supabase becomes the archive.

Env:
  MMS_TOKEN                  accessToken cookie from merchant.shoalter.com
  SUPABASE_URL               https://<ref>.supabase.co
  SUPABASE_SERVICE_ROLE_KEY  service_role key (never ship this to the browser)

Usage:
  python mms_to_supabase.py --store B0812001 --days 12
  python mms_to_supabase.py --store B0812001 --start 2026-08-22 --end 2026-09-02
"""
import argparse, io, os, sys
from collections import defaultdict
from datetime import date, datetime, timedelta

import requests
from openpyxl import load_workbook

MMS = "https://merchant-web.shoalter.com"
BATCH = 10          # dates per metadata call
CHUNK = 500         # rows per Supabase upsert


def daterange(a, b):
    d = a
    while d <= b:
        yield d
        d += timedelta(days=1)


def report_files(s, bu, store, dates):
    out = []
    for i in range(0, len(dates), BATCH):
        chunk = ",".join(d.strftime("%Y%m%d") for d in dates[i:i + BATCH])
        r = s.get(f"{MMS}/order/{bu}/{store}/DAILY/report",
                  params={"dates": chunk}, timeout=60)
        if r.status_code != 200 or r.json().get("status", {}).get("code") != "success":
            raise SystemExit(f"MMS metadata failed ({r.status_code}). "
                             f"If it says 'Unable to get information from mms-user', "
                             f"your MMS_TOKEN expired - grab a fresh one.\n{r.text[:300]}")
        out += r.json()["data"]["data"]
    return out


def read_report(s, bu, filename):
    r = s.get(f"{MMS}/order/{bu}/DAILY/{filename}/downloadreport", timeout=120)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if any(str(c).strip() == "SKU ID" for c in row if c is not None):
                header = [str(c).replace("\n", " ").strip() if c else "" for c in row]
            continue
        rec = dict(zip(header, row))
        if rec.get("SKU ID"):
            yield rec
    wb.close()


def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)[:10]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--store", required=True)
    p.add_argument("--bu", default="HKTV")
    p.add_argument("--start")
    p.add_argument("--end")
    p.add_argument("--days", type=int, default=12)
    p.add_argument("--dry-run", action="store_true")
    a = p.parse_args()

    token = os.environ.get("MMS_TOKEN", "")
    sb_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    sb_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not token:
        sys.exit("Missing MMS_TOKEN")
    if not a.dry_run and not (sb_url and sb_key):
        sys.exit("Missing SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")

    if a.start:
        d0 = datetime.strptime(a.start, "%Y-%m-%d").date()
        d1 = datetime.strptime(a.end, "%Y-%m-%d").date()
    else:
        d1 = date.today()
        d0 = d1 - timedelta(days=a.days - 1)

    mms = requests.Session()
    mms.headers.update({"Authorization": "Bearer " + token})

    sb = requests.Session()
    sb.headers.update({"apikey": sb_key, "Authorization": "Bearer " + sb_key,
                       "Content-Type": "application/json"})

    dates = list(daterange(d0, d1))
    files = report_files(mms, a.bu, a.store, dates)
    have = {f.split("_")[-1][:8] for f in files}
    gap = [d.strftime("%Y%m%d") for d in dates if d.strftime("%Y%m%d") not in have]
    if gap:
        print(f"WARNING: MMS has no report for {','.join(gap)} "
              f"(retention is ~12 days; those days are gone unless already archived)",
              file=sys.stderr)

    grand = 0
    for fn in files:
        order_date = f"{fn.split('_')[-1][:4]}-{fn.split('_')[-1][4:6]}-{fn.split('_')[-1][6:8]}"
        agg = defaultdict(lambda: {"qty": 0.0, "amount": 0.0, "lines": 0,
                                   "zh": None, "en": None})
        for rec in read_report(mms, a.bu, fn):
            k = (as_date(rec.get("Delivery Date")), str(rec["SKU ID"]))
            r = agg[k]
            r["qty"] += float(rec.get("Qty (Q)") or 0)
            r["amount"] += float(rec.get("Total Q * U - D = T") or 0)
            r["lines"] += 1
            r["zh"] = rec.get("SKU Name (Chinese)") or r["zh"]
            r["en"] = rec.get("SKU Name (English)") or r["en"]

        rows = [{"store_code": a.store, "order_date": order_date,
                 "delivery_date": dd, "sku_id": sku,
                 "sku_name_zh": v["zh"], "sku_name_en": v["en"],
                 "qty": round(v["qty"], 3), "amount": round(v["amount"], 2),
                 "lines": v["lines"]}
                for (dd, sku), v in agg.items()]
        grand += len(rows)
        print(f"  {order_date}: {len(rows)} rows", file=sys.stderr)
        if a.dry_run:
            continue

        # authoritative replace for this (store, order_date)
        r = sb.delete(f"{sb_url}/rest/v1/mms_sku_daily",
                      params={"store_code": f"eq.{a.store}",
                              "order_date": f"eq.{order_date}"}, timeout=60)
        r.raise_for_status()
        for i in range(0, len(rows), CHUNK):
            r = sb.post(f"{sb_url}/rest/v1/mms_sku_daily",
                        headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
                        json=rows[i:i + CHUNK], timeout=120)
            if r.status_code >= 300:
                sys.exit(f"Supabase write failed {r.status_code}: {r.text[:400]}")

    print(f"{len(files)} day-file(s), {grand} aggregate rows"
          f"{' (dry run, nothing written)' if a.dry_run else ' upserted'}",
          file=sys.stderr)


if __name__ == "__main__":
    main()
